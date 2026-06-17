from dataclasses import dataclass
from typing import Any, Callable
from openpyxl import load_workbook

from custom.hpgds.constants import (EFFORT_SUMMARY_SHEET, STUDIO_SUPPORT_SUMMARY_SHEET)
from custom.hpgds.calculators.employee_forecast_builder import build_employee_forecast
from custom.hpgds.reports.executive_summary import build_executive_summary
from custom.hpgds.reports.studio_support_breakdown import build_studio_support_breakdown
from custom.hpgds.reports.validation_summary import build_validation_summary
from custom.hpgds.reports.hpgds_forecast import build_hpgds_forecast
from custom.hpgds.reports.final_summary import build_final_summary
from custom.hpgds.reports.forecast_studio_breakdown import build_forecast_studio_breakdown
from custom.hpgds.reports.final_studio_breakdown import build_final_studio_breakdown


@dataclass
class ReportDefinition:
    sheet_name: str
    data: Any
    writer: Callable


def write_summary(worksheet, summary, start_row=2):
    """
    Generic writer for simple Category -> Hours reports.
    """

    row = start_row

    for category, hours in summary.items():
        worksheet.cell(row=row, column=1, value=category)
        worksheet.cell(row=row, column=2, value=hours)
        row += 1

    # Clear unused rows
    for clear_row in range(row, 101):
        worksheet.cell(row=clear_row, column=1, value=None)
        worksheet.cell(row=clear_row, column=2, value=None)


def write_reports(workbook, reports):
    """
    Generic report writer.
    """

    for report in reports:
        if report.sheet_name not in workbook.sheetnames:
            raise ValueError(f"Worksheet '{report.sheet_name}' not found in template.")

        worksheet = workbook[report.sheet_name]
        report.writer(worksheet, report.data )


def generate_workbook(
    worklogs,
    validation_issues,
    studio_groups_config,
    template_file,
    output_file,
    forecast_request=None,
):
    workbook = load_workbook(template_file)

    # =====================================================
    # Build Base Data
    # =====================================================

    actual_summary = build_executive_summary(worklogs, studio_groups_config)
    studio_breakdown = build_studio_support_breakdown(worklogs, studio_groups_config)
    validation_summary = build_validation_summary(validation_issues)

    # =====================================================
    # Forecast Data (Optional)
    # =====================================================

    if forecast_request is not None:

        employee_forecast = build_employee_forecast(worklogs, forecast_request)
        forecast_summary = build_hpgds_forecast(worklogs, studio_groups_config, employee_forecast)
        final_summary = build_final_summary(actual_summary, forecast_summary)
        forecast_breakdown = build_forecast_studio_breakdown(employee_forecast, studio_breakdown)
        final_breakdown = build_final_studio_breakdown(studio_breakdown, forecast_breakdown)

    else:

        final_summary = actual_summary
        final_breakdown = studio_breakdown

    # =====================================================
    # Report Definitions
    # =====================================================

    reports = [
        ReportDefinition(
            sheet_name=EFFORT_SUMMARY_SHEET,
            data=final_summary,
            writer=write_summary,
        ),
        ReportDefinition(
            sheet_name=STUDIO_SUPPORT_SUMMARY_SHEET,
            data=final_breakdown,
            writer=write_summary,
        ),
        ReportDefinition(
            sheet_name="Validation Summary",
            data=validation_summary,
            writer=write_summary,
        ),
    ]

    write_reports(workbook, reports)
    workbook.save(output_file)
