from openpyxl import Workbook, load_workbook

from worklog_analytics.reports.excel_renderer import write_reports


def write_workbook(reports, output_file, template_file=None,):
    if template_file:
        workbook = load_workbook(template_file)
    else:
        workbook = Workbook()

    write_reports(workbook, reports)
    workbook.save(output_file)
